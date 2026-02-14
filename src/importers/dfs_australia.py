from __future__ import annotations

"""Import DFS Australia SuperCoach spreadsheet data into the database.

The spreadsheet (typically named afl-supercoach-YYYY.xlsx) contains ~780 players
with salaries, ownership %, advanced metrics (CBA%, PPM), historical averages,
and state league stats. The importer:
  1. Reads the Excel file with openpyxl
  2. Upserts Player records (matching by champion_data_id or name)
  3. Upserts DfsPlayerStats records (one per player per season)
"""

import logging
from pathlib import Path
from typing import Optional

import openpyxl
from sqlalchemy import select

from src.models.database import DfsPlayerStats, Player, get_session, init_db

logger = logging.getLogger(__name__)

# Column indices in the DFS Australia spreadsheet (0-based, from row 9 header)
COL_PLAYER = 0
COL_CD_ID = 1
COL_LINK = 2
COL_TEAM = 3
COL_SALARY = 4
COL_OWNED = 5
COL_POSITION = 6
COL_GAMES = 7
COL_SC_AVG = 8
COL_SC_MAX = 9
COL_SC_100 = 10
COL_SC_120 = 11
COL_CBA_PCT = 12
COL_PPM = 13
COL_REG = 14
COL_L5 = 15
COL_FIN = 16
COL_PREV_YEAR = 17
COL_PREV_2YR = 18
# State league stats
COL_VFL_GAMES = 19
COL_VFL_FP = 20
COL_VFL_MAX = 21
COL_VFL_100 = 22
COL_WAFL_GAMES = 23
COL_WAFL_FP = 24
COL_WAFL_MAX = 25
COL_WAFL_100 = 26
COL_SANFL_GAMES = 27
COL_SANFL_FP = 28
COL_SANFL_MAX = 29
COL_SANFL_100 = 30
COL_SANFL_U18_GAMES = 31
COL_SANFL_U18_FP = 32
COL_SANFL_U18_MAX = 33
COL_SANFL_U18_100 = 34
COL_CTL_GAMES = 35
COL_CTL_FP = 36
COL_CTL_MAX = 37
COL_CTL_100 = 38

# DFS Australia team abbreviations -> our standard team names
TEAM_MAP = {
    "ADE": "Adelaide",
    "BRL": "Brisbane",
    "CAR": "Carlton",
    "COL": "Collingwood",
    "ESS": "Essendon",
    "FRE": "Fremantle",
    "GCS": "Gold Coast",
    "GEE": "Geelong",
    "GWS": "GWS",
    "HAW": "Hawthorn",
    "MEL": "Melbourne",
    "NTH": "North Melbourne",
    "PTA": "Port Adelaide",
    "RIC": "Richmond",
    "STK": "St Kilda",
    "SYD": "Sydney",
    "WBD": "Western Bulldogs",
    "WCE": "West Coast",
}

# Header row is row 9 (1-indexed) — data starts at row 10
HEADER_ROW = 9
DATA_START_ROW = 10


def _safe_int(val: object) -> Optional[int]:
    """Safely convert a value to int, returning None for non-numeric."""
    if val is None:
        return None
    try:
        return int(val)
    except (ValueError, TypeError):
        return None


def _safe_float(val: object) -> Optional[float]:
    """Safely convert a value to float, returning None for non-numeric."""
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _normalise_position(raw: Optional[str]) -> Optional[str]:
    """Normalise DFS position labels to match our DB conventions.

    DFS uses: MID, DEF, FWD, RUC, MID/FWD, DEF/MID, etc.
    We keep them as-is since they contain useful dual-position info.
    """
    if not raw:
        return None
    return raw.strip().upper()


def _find_or_create_player(
    session: object,
    name: str,
    champion_data_id: str,
    team_abbr: str,
    position: Optional[str],
) -> Player:
    """Find a player by champion_data_id or name, or create a new one."""
    # 1. Try by champion_data_id first (most reliable)
    player = session.execute(  # type: ignore[union-attr]
        select(Player).where(Player.champion_data_id == champion_data_id)
    ).scalar_one_or_none()
    if player:
        # Update fields that may have changed
        team_name = TEAM_MAP.get(team_abbr, team_abbr)
        player.team = team_name
        if position:
            player.position = position
        return player

    # 2. Try exact name match first, then fuzzy (for players from FootyWire)
    team_name = TEAM_MAP.get(team_abbr, team_abbr)

    # Exact match — may return multiple (player traded between teams)
    candidates = session.execute(  # type: ignore[union-attr]
        select(Player).where(Player.name == name)
    ).scalars().all()

    if not candidates:
        # Fuzzy match
        candidates = session.execute(  # type: ignore[union-attr]
            select(Player).where(Player.name.ilike(f"%{name}%"))
        ).scalars().all()

    if candidates:
        # Pick the best match — prefer same team
        player = None
        for c in candidates:
            if c.team == team_name:
                player = c
                break
        if player is None:
            player = candidates[0]

        # Link champion_data_id and update
        player.champion_data_id = champion_data_id
        player.team = team_name
        if position:
            player.position = position
        return player

    # 3. Create new player
    team_name = TEAM_MAP.get(team_abbr, team_abbr)
    player = Player(
        name=name,
        team=team_name,
        position=position,
        champion_data_id=champion_data_id,
    )
    session.add(player)  # type: ignore[union-attr]
    session.flush()  # type: ignore[union-attr]
    return player


def import_dfs_spreadsheet(file_path: str, season: int = 2026) -> int:
    """Import a DFS Australia Excel spreadsheet into the database.

    Args:
        file_path: Path to the .xlsx file
        season: Season year this data is for (default: 2026)

    Returns:
        Number of players imported/updated
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    init_db()
    session = get_session()

    try:
        wb = openpyxl.load_workbook(str(path), read_only=True)
        # Use the first sheet
        ws = wb[wb.sheetnames[0]]

        count = 0
        row_idx = 0

        for row in ws.iter_rows(values_only=True):
            row_idx += 1

            # Skip header rows (data starts at row 10)
            if row_idx < DATA_START_ROW:
                continue

            # Check we have a valid player row
            player_name = row[COL_PLAYER] if len(row) > COL_PLAYER else None
            cd_id = row[COL_CD_ID] if len(row) > COL_CD_ID else None

            if not player_name or not cd_id:
                continue

            # Skip if it looks like a header/section row
            if not isinstance(player_name, str) or not isinstance(cd_id, str):
                continue
            if not cd_id.startswith("CD_I"):
                continue

            player_name = player_name.strip()
            cd_id = cd_id.strip()
            team_abbr = str(row[COL_TEAM]).strip() if row[COL_TEAM] else "Unknown"
            position = _normalise_position(str(row[COL_POSITION]) if row[COL_POSITION] else None)

            # Find or create the player
            player = _find_or_create_player(session, player_name, cd_id, team_abbr, position)

            # Upsert DfsPlayerStats
            dfs_stat = session.execute(  # type: ignore[union-attr]
                select(DfsPlayerStats).where(
                    DfsPlayerStats.player_id == player.id,
                    DfsPlayerStats.season == season,
                )
            ).scalar_one_or_none()

            if dfs_stat is None:
                dfs_stat = DfsPlayerStats(player_id=player.id, season=season)
                session.add(dfs_stat)  # type: ignore[union-attr]

            # Populate stats from the row
            dfs_stat.salary = _safe_int(row[COL_SALARY])
            dfs_stat.ownership_pct = _safe_float(row[COL_OWNED])
            dfs_stat.games_played = _safe_int(row[COL_GAMES])
            dfs_stat.sc_avg = _safe_float(row[COL_SC_AVG])
            dfs_stat.sc_max = _safe_int(row[COL_SC_MAX])
            dfs_stat.scores_100_plus = _safe_int(row[COL_SC_100])
            dfs_stat.scores_120_plus = _safe_int(row[COL_SC_120])
            dfs_stat.cba_pct = _safe_float(row[COL_CBA_PCT])
            dfs_stat.points_per_minute = _safe_float(row[COL_PPM])
            dfs_stat.reg_avg = _safe_float(row[COL_REG])
            dfs_stat.last_5_avg = _safe_float(row[COL_L5])
            dfs_stat.finals_avg = _safe_float(row[COL_FIN])
            dfs_stat.prev_year_avg = _safe_float(row[COL_PREV_YEAR])
            dfs_stat.prev_2yr_avg = _safe_float(row[COL_PREV_2YR])

            # State league stats
            dfs_stat.vfl_games = _safe_int(row[COL_VFL_GAMES]) if len(row) > COL_VFL_GAMES else None
            dfs_stat.vfl_avg = _safe_float(row[COL_VFL_FP]) if len(row) > COL_VFL_FP else None
            dfs_stat.vfl_max = _safe_int(row[COL_VFL_MAX]) if len(row) > COL_VFL_MAX else None
            dfs_stat.vfl_100_plus = _safe_int(row[COL_VFL_100]) if len(row) > COL_VFL_100 else None

            dfs_stat.wafl_games = _safe_int(row[COL_WAFL_GAMES]) if len(row) > COL_WAFL_GAMES else None
            dfs_stat.wafl_avg = _safe_float(row[COL_WAFL_FP]) if len(row) > COL_WAFL_FP else None
            dfs_stat.wafl_max = _safe_int(row[COL_WAFL_MAX]) if len(row) > COL_WAFL_MAX else None
            dfs_stat.wafl_100_plus = _safe_int(row[COL_WAFL_100]) if len(row) > COL_WAFL_100 else None

            dfs_stat.sanfl_games = _safe_int(row[COL_SANFL_GAMES]) if len(row) > COL_SANFL_GAMES else None
            dfs_stat.sanfl_avg = _safe_float(row[COL_SANFL_FP]) if len(row) > COL_SANFL_FP else None
            dfs_stat.sanfl_max = _safe_int(row[COL_SANFL_MAX]) if len(row) > COL_SANFL_MAX else None
            dfs_stat.sanfl_100_plus = _safe_int(row[COL_SANFL_100]) if len(row) > COL_SANFL_100 else None

            dfs_stat.sanfl_u18_games = (
                _safe_int(row[COL_SANFL_U18_GAMES]) if len(row) > COL_SANFL_U18_GAMES else None
            )
            dfs_stat.sanfl_u18_avg = (
                _safe_float(row[COL_SANFL_U18_FP]) if len(row) > COL_SANFL_U18_FP else None
            )
            dfs_stat.sanfl_u18_max = (
                _safe_int(row[COL_SANFL_U18_MAX]) if len(row) > COL_SANFL_U18_MAX else None
            )
            dfs_stat.sanfl_u18_100_plus = (
                _safe_int(row[COL_SANFL_U18_100]) if len(row) > COL_SANFL_U18_100 else None
            )

            dfs_stat.ctl_games = _safe_int(row[COL_CTL_GAMES]) if len(row) > COL_CTL_GAMES else None
            dfs_stat.ctl_avg = _safe_float(row[COL_CTL_FP]) if len(row) > COL_CTL_FP else None
            dfs_stat.ctl_max = _safe_int(row[COL_CTL_MAX]) if len(row) > COL_CTL_MAX else None
            dfs_stat.ctl_100_plus = (
                _safe_int(row[COL_CTL_100]) if len(row) > COL_CTL_100 else None
            )

            count += 1

            if count % 100 == 0:
                logger.info("Imported %d players...", count)

        session.commit()
        wb.close()

        logger.info("DFS Australia import complete: %d players", count)
        return count

    except Exception:
        session.rollback()
        raise
    finally:
        session.close()
